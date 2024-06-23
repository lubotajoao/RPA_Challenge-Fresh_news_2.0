import os
import traceback
import openpyxl

log_path = ""
log_file = ""
current_folder = ""
excel_report_file = ""


def create_excel_report_file(report_path: str, current_date: str, project_name: str) -> str:
    global excel_report_file

    try:
        excel_report_file = f"{report_path}/Report_{project_name}_{current_date}.xlsx"

        if not os.path.isfile(excel_report_file):
            workbook = openpyxl.Workbook()
            workbook.save(excel_report_file)

            workbook = openpyxl.load_workbook(filename=excel_report_file)
            sheet = workbook.active
            sheet.title = "Fresh News 2.0 Report"
            workbook.save(filename=excel_report_file)

            workbook = openpyxl.load_workbook(filename=excel_report_file)
            sheet = workbook["Fresh News 2.0 Report"]
            sheet["A1"] = "Title"
            sheet["B1"] = "Date"
            sheet["C1"] = "Description"
            sheet["D1"] = "Picture Filename"
            sheet["E1"] = "Count Of Search Phrase"
            sheet["F1"] = "Picture News Path"

            workbook.save(filename=excel_report_file)
            workbook.close()

        return excel_report_file

    except FileNotFoundError as e:
        print(f'An error occurred while creating the Excel report file[{excel_report_file}]:\n[{e}][\n{traceback.format_exc()}]')

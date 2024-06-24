import traceback
from openpyxl import load_workbook

from helper.csv_log import csv_log

module_name = "excel_set_report_info.py"


def get_current_index(report_file):
    workbook = load_workbook(filename=report_file, data_only=True)
    sheet = workbook["Fresh News 2.0 Report"]
    rows = sheet.iter_rows(min_row=2, max_col=3)
    index = 2
    for a, b, c in rows:
        if a.value is not None and b.value is not None and c.value is not None:
            index += 1
            continue
    workbook.save(filename=report_file)
    workbook.close()
    return index


def excel_set_report_info(log_file: str, project_name: str, report_file: str, title: str, date: str, description: str, picture_filename: str, count_search_phrase: str,
                          picture_file_path: str) -> None:
    global module_name
    try:
        index = get_current_index(report_file=report_file)
        workbook = load_workbook(filename=report_file)

        sheet = workbook["Fresh News 2.0 Report"]
        sheet[f"A{index}"] = title
        sheet[f"B{index}"] = date
        sheet[f"C{index}"] = description
        sheet[f"D{index}"] = picture_filename
        sheet[f"E{index}"] = count_search_phrase
        sheet[f"F{index}"] = picture_file_path

        workbook.save(filename=report_file)
        workbook.close()

        status = "Info"
        description = f'The info of [{title}] inserted successfully in Excel report file.'
        csv_log(file=log_file, project_name=project_name, task=module_name, status=status, description=description)

    except Exception:
        status = "Error"
        description = f'An error occurred in excel_set_report_info Module: [{traceback.format_exc()}].'
        csv_log(file=log_file, project_name=project_name, task=module_name, status=status, description=description)
